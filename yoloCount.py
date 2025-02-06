import subprocess
import openpyxl
import json
import os

# mockData = '{"7": {"car": 2}, "4": {"car": 18}, "3": {"car": 1}, "5": {"truck": 1}}'

def run(folder_path):
    # Call the second Python script and capture its output
    file_list = os.listdir(folder_path)

    for file_name in file_list:
        full_path = os.path.join(folder_path, file_name)
        if os.path.isdir(full_path):
            run(full_path)
        elif file_name.endswith('.dav') or file_name.endswith('.mp4'):
            params = [
            '--video_file_path', full_path
            ]
            result = subprocess.run(['yolovenv/Scripts/python', 'app.py'] + params, capture_output=True, text=True)
            output = result.stdout.strip()
            if "." in output:
                output = output.replace('\n', ' ').split(".")[2].strip()
            if output and output.startswith('{') and output.endswith('}'):
                variables = json.loads(output)
                buildExcel(variables, file_name)
            else:
                print(f"Invalid or empty output for file: {file_name}")

def buildExcel(variables, file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    startRow =2
    headers = set()

    ws.cell(row=startRow, column=1, value='Movimiento')

    for key,value in variables.items():
        headers.update(value.keys())
    headers = list(headers)
    
    for col_num, header in enumerate(headers, start=2):
        ws.cell(row=startRow, column=col_num, value=header)

    startRow += 1
    for property_name, value in variables.items():
        ws.cell(row=startRow, column=1, value=property_name)
        for col_num, header in enumerate(headers, start=2):
            ws.cell(row=startRow, column=col_num, value=value.get(header, 0))
        startRow += 1

    ws.cell(row=startRow, column=1, value='Grand Total')
    for col_num, header in enumerate(headers, start=2):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sum_formula = f"=SUM({col_letter}3:{col_letter}{startRow-1})"
            ws.cell(row=startRow, column=col_num, value=sum_formula)

    wb.save(f"{file_name}.xlsx")


if __name__ == "__main__":
    folderPath = "./data"
    run(folderPath)